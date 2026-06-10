"""
Content outline cog: scans a channel for inspiration links and exports
a content creation outline as an xlsx file.
"""

import io
import re
import logging
from datetime import datetime
from typing import Optional
from urllib.parse import urlparse

import discord
from discord import app_commands
from discord.ext import commands

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

URL_REGEX = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)

PLATFORM_MAP = {
    "instagram.com": "Instagram",
    "instagr.am": "Instagram",
    "tiktok.com": "TikTok",
    "vm.tiktok.com": "TikTok",
    "youtube.com": "YouTube",
    "youtu.be": "YouTube",
    "twitter.com": "Twitter/X",
    "x.com": "Twitter/X",
    "facebook.com": "Facebook",
    "fb.watch": "Facebook",
    "linkedin.com": "LinkedIn",
    "pinterest.com": "Pinterest",
    "pin.it": "Pinterest",
    "reddit.com": "Reddit",
    "threads.net": "Threads",
    "vimeo.com": "Vimeo",
}

OUTLINE_HEADERS = [
    "#",
    "Inspiration Link",
    "Platform",
    "Shared By",
    "Shared On",
    "Message Context",
    "Hook / Concept",
    "Target Audience",
    "Format (Reel / Post / Story / Video)",
    "Caption Draft",
    "Call To Action",
    "Hashtags",
    "Assigned To",
    "Status",
    "Notes",
]


def _platform_from_url(url: str) -> str:
    try:
        host = urlparse(url).netloc.lower()
        if host.startswith("www."):
            host = host[4:]
        for key, name in PLATFORM_MAP.items():
            if host == key or host.endswith("." + key):
                return name
        return host or "Unknown"
    except Exception:
        return "Unknown"


def _build_workbook(channel_name: str, rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Outline"

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTLINE_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"Content Creation Outline — #{channel_name}")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.cell(
        row=2,
        column=1,
        value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} • {len(rows)} inspiration link(s)",
    ).font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(OUTLINE_HEADERS))

    header_row = 4
    for col, header in enumerate(OUTLINE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    for idx, row in enumerate(rows, start=1):
        r = header_row + idx
        values = [
            idx,
            row["url"],
            row["platform"],
            row["author"],
            row["timestamp"],
            row["context"],
            "", "", "", "", "", "", "", "Idea", "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.alignment = wrap
            if col == 2 and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[r].height = 60

    widths = [5, 50, 14, 18, 18, 40, 32, 22, 22, 40, 22, 28, 18, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class ContentOutlineCog(commands.Cog):
    """Cog providing the /content-outline command."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(
        name="content-outline",
        description="Scan a channel for inspiration links and export a content outline xlsx",
    )
    @app_commands.describe(
        channel="The channel to scan for inspiration links",
        limit="Maximum number of messages to scan (default 1000, max 5000)",
    )
    async def content_outline(
        self,
        interaction: discord.Interaction,
        channel: discord.TextChannel,
        limit: Optional[int] = 1000,
    ):
        if not interaction.user.guild_permissions.send_messages:
            await interaction.response.send_message(
                "❌ You don't have permission to use this command.", ephemeral=True
            )
            return

        perms = channel.permissions_for(interaction.guild.me)
        if not (perms.read_messages and perms.read_message_history):
            await interaction.response.send_message(
                f"❌ I don't have permission to read history in {channel.mention}.",
                ephemeral=True,
            )
            return

        scan_limit = max(1, min(limit or 1000, 5000))

        await interaction.response.defer(thinking=True)

        rows: list[dict] = []
        seen_urls: set[str] = set()

        try:
            async for message in channel.history(limit=scan_limit, oldest_first=True):
                if message.author.bot:
                    continue
                content = message.content or ""
                urls = URL_REGEX.findall(content)
                for raw_url in urls:
                    url = raw_url.rstrip(").,;:!?\"'>")
                    if url in seen_urls:
                        continue
                    seen_urls.add(url)

                    context = URL_REGEX.sub("", content).strip()
                    context = re.sub(r"\s+", " ", context)
                    if len(context) > 300:
                        context = context[:297] + "..."

                    rows.append({
                        "url": url,
                        "platform": _platform_from_url(url),
                        "author": message.author.display_name,
                        "timestamp": message.created_at.strftime("%Y-%m-%d %H:%M"),
                        "context": context,
                    })
        except discord.Forbidden:
            await interaction.followup.send(
                f"❌ Missing permissions to read {channel.mention}.", ephemeral=True
            )
            return
        except Exception as e:
            logger.error(f"Error scanning channel {channel.id}: {e}", exc_info=True)
            await interaction.followup.send(
                f"❌ Error scanning channel: {e}", ephemeral=True
            )
            return

        if not rows:
            await interaction.followup.send(
                f"📭 No links found in the last {scan_limit} messages of {channel.mention}.",
                ephemeral=True,
            )
            return

        try:
            xlsx_bytes = _build_workbook(channel.name, rows)
        except Exception as e:
            logger.error(f"Error building xlsx: {e}", exc_info=True)
            await interaction.followup.send(
                f"❌ Failed to build xlsx file: {e}", ephemeral=True
            )
            return

        filename = f"content-outline-{channel.name}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        file = discord.File(io.BytesIO(xlsx_bytes), filename=filename)

        embed = discord.Embed(
            title="📋 Content Outline Generated",
            description=(
                f"Pulled **{len(rows)}** inspiration link(s) from {channel.mention}\n"
                f"Scanned up to **{scan_limit}** messages."
            ),
            color=0x2E75B6,
        )
        embed.set_footer(text=f"Requested by {interaction.user.display_name}")
        embed.timestamp = datetime.now()

        await interaction.followup.send(embed=embed, file=file)


async def setup(bot: commands.Bot):
    await bot.add_cog(ContentOutlineCog(bot))
